from Train_Validation import *
from load_data import *
from save_load_model import *
from auto_encoder_MLP import *
from feature_selection_method import *
from openpyxl import load_workbook
import os
import argparse
import numpy as np
import torch.utils.data
import torch.nn as nn
import torch.optim as optim
from os.path import join
from os import listdir
from torch.utils.data import DataLoader


def flatten_one(length, img):
    one_line = np.zeros((1, int(length)))
    position = 0
    for i in range(img.shape[0]):  # column
        for j in range(i + 1, img.shape[1]):  # row
            one_line[0, position] = img[j, i]
            position = position + 1
    return one_line
########################################### Load Data ###############################################
#####################################################################################################
#####################################################################################################
root_path = "Input the full path of \\download_data\\all  (all folder will be created by Data_download_prepare\\fetch_data.py)"
cross_validation_path = 'Input the full path of \\download_data\\cross validation  (cross validation folder will be created by Data_download_prepare\\fetch_data.py)'
label_dir = 'Input the full path of \\download_data\\measures  (measures folder will be created by Data_download_prepare\\fetch_data.py)'


k_fold = 5
number_samples = 871
image_size = [200, 200]   # this is the size of cc200 atlas.
batch_size_train = 511
batch_size_validation = 167
batch_size_test = 167
new_number_features = 5000


length = image_size[0] * (image_size[1]-1) / 2
group_images = np.zeros((number_samples, int(length)))
for i in range(number_samples):
    img_name = str(i) + '.mat'
    img = scio.loadmat(os.path.join(root_path, img_name))
    img = img['connectivity']
    img = flatten_one(length, img)
    group_images[i, :] = img

################################################ selectors ###########################################
label_name = 'ABIDE_label_871.mat'
label = scio.loadmat(os.path.join(label_dir, label_name))
label = label['label']
label = label[0]
dist_selectors = {}
for i in range(k_fold):
    train_path = cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'train' + '\\'
    filenames = listdir(train_path)
    train_ind = []
    for j in filenames:
        index = int(j[:-4])
        train_ind.append(index)

    selector = feature_selection(group_images, label, train_ind, new_number_features)
    dist_selectors[str(i+1)] = selector


dist_train = {}
dist_validation = {}
dist_test = {}
for i in range(k_fold):
    fold_list = []
    for j in range(k_fold):
        fold_list.append('set' + str(j))
    del fold_list[i]
    dist_validation[str(i+1)] = TrainDataset_DNN(
                            root_dir=join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'validation'),
                            label_dir = join(label_dir),
                            selector = dist_selectors[str(i+1)])

    dist_train[str(i+1)] = TrainDataset_DNN(
                            root_dir=join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'train'),
                            label_dir=join(label_dir),
                            selector = dist_selectors[str(i+1)])

    dist_test[str(i + 1)] = TrainDataset_DNN(
                            root_dir=join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'test'),
                            label_dir=join(label_dir),
                            selector = dist_selectors[str(i+1)])

# ########################################### Training ################################################
# #####################################################################################################
# #####################################################################################################
# #
parser = argparse.ArgumentParser()
parser.add_argument('--ngpu', type=int, default=1)
parser.add_argument('--nEpochs', type=int, default=150)
parser.add_argument('--weight-decay', '--wd', default=1e-8, type=float,
                    metavar='W', help='weight decay (default: 1e-8)')
parser.add_argument('--no-cuda', action='store_true')
parser.add_argument('--seed', type=int, default=1)
parser.add_argument('--opt', type=str, default='adam',
                    choices=('sgd', 'adam', 'rmsprop'))
parser.add_argument('--auto_in', type=int, default=5000)
parser.add_argument('--auto_hid_1', type=int, default=200)
parser.add_argument('--auto_hid_2', type=int, default=150)
parser.add_argument('--auto_hid_3', type=int, default=200)

parser.add_argument('--MLP_1', type=int, default=300)
parser.add_argument('--MLP_2', type=int, default=16)
parser.add_argument('--MLP_out', type=int, default=2)
parser.add_argument('--dropout_rate', type=int, default=0.1)


args, unknown = parser.parse_known_args()
args.cuda = not args.no_cuda and torch.cuda.is_available()
weight_decay = args.weight_decay
torch.manual_seed(args.seed)
if args.cuda:
    torch.cuda.manual_seed(args.seed)


print("build auto-encoder")
model = Auto_encoder_MLP(in_c=args.auto_in, auto_1=args.auto_hid_1, auto_2=args.auto_hid_2, auto_3=args.auto_hid_3, MLP_1=args.MLP_1, MLP_2=args.MLP_2, MLP_out=args.MLP_out, dropout_rate=args.dropout_rate)
gpu_ids = range(args.ngpu)
model = nn.parallel.DataParallel(model, device_ids=gpu_ids)
model.apply(weights_init)

train = train_DNN
test = test_DNN


print('  + Number of params: {}'.format(
    sum([p.data.nelement() for p in model.parameters()])))
if args.cuda:
    model = model.cuda()


if args.opt == 'sgd':
    optimizer = optim.SGD(model.parameters(),lr=1e-5, momentum=0.8)
elif args.opt == 'adam':
    optimizer = optim.Adam(model.parameters())
elif args.opt == 'rmsprop':
    optimizer = optim.RMSprop(model.parameters(), lr=1e-3, weight_decay=weight_decay)


save_path = 'Choose a path where you want to store the models'
save_model(model,save_path, 'initial_model')

###################################### excel #########################################################
excel_root = "Choose a path where you want to record the accuracy and loss during the training validation and testing processes"
excel_name = 'Auto.xlsx'

workbook=load_workbook(os.path.join(excel_root, excel_name))
Sheet1 = workbook['Sheet1']

rows_each_section = 20
row_train_accuracy = 1
row_train_loss = 9
position = [row_train_accuracy, row_train_loss, rows_each_section]

Sheet1.cell(row=position[0], column=1, value='Train_accuracy')
Sheet1.cell(row=position[1], column=1, value='Train_loss')
Sheet1.cell(row=position[0] + position[2], column=1, value='Validation_accuracy')
Sheet1.cell(row=position[1] + position[2], column=1, value='Validation_loss')
Sheet1.cell(row=position[0] + position[2] * 2, column=1, value='Test_accuracy')
Sheet1.cell(row=position[1] + position[2] * 2, column=1, value='Test_loss')
Sheet1.cell(row=position[0] + position[2] * 3, column=1, value='Sensitivity')
Sheet1.cell(row=position[1] + position[2] * 3, column=1, value='Specificity')

################################################# Training ###################################################
for fold in range(1, k_fold+1):
    trainLoader = DataLoader(
        dist_train[str(fold)],
        shuffle=True,
        batch_size=batch_size_train,
        drop_last=True
    )

    validationLoader = DataLoader(
        dist_validation[str(fold)],
        shuffle=True,
        batch_size=batch_size_validation,
        drop_last=True
    )

    testLoader = DataLoader(
        dist_test[str(fold)],
        shuffle=True,
        batch_size=batch_size_test,
        drop_last=True
    )

    Sheet1.cell(row=fold + position[0], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2] * 2, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2] * 2, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2] * 3, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2] * 3, column=1, value='model_' + str(fold))


    for epoch in range(1, args.nEpochs + 1):
        model.apply(weights_init)
        if epoch == 1:
            model = load_model(model, save_path, 'initial_model')
            print('model: ' + str(epoch) + '_' + str(fold))
        else:
            model = load_model(model, save_path, str(epoch - 1) + '_' + str(fold))
            print('model: ' + str(epoch) + '_' + str(fold))

        adjust_opt_auto(args.opt, optimizer, epoch)
        train(args, epoch, model, trainLoader, validationLoader, testLoader, optimizer, fold, save_path, Sheet1, position)

workbook.save(os.path.join(excel_root, excel_name))
