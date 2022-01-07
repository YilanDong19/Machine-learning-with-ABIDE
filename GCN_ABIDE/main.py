from GCN import *
from Train_Validation import *
from prepare_ABIDE import *
from save_load_model import *
from openpyxl import load_workbook
import os
import argparse
import numpy as np
import torch.utils.data
import torch.nn as nn
import torch.optim as optim
from os import listdir
# %matplotlib inline



########################################### Load Data ###############################################
#####################################################################################################
#####################################################################################################
root_path = "Input the full path of \\download_data\\all  (all folder will be created by Data_download_prepare\\fetch_data.py)"
cross_validation_path = 'Input the full path of \\download_data\\cross validation  (cross validation folder will be created by Data_download_prepare\\fetch_data.py)'
label_dir = 'Input the full path of \\download_data\\measures  (measures folder will be created by Data_download_prepare\\fetch_data.py)'

k_fold = 5
number_samples = 871
######################################### extract 5000 features #############################
new_number_features =5000
all_networks = []
subject_IDs = []
for i in range(number_samples):
    subject_IDs.append(str(i))
    img_name = str(i) + '.mat'
    img = scio.loadmat(os.path.join(root_path, img_name))
    img = img['connectivity']
    all_networks.append(img)

idx = np.triu_indices_from(all_networks[0], 1)
norm_networks = [np.arctanh(mat) for mat in all_networks]
vec_networks = [mat[idx] for mat in all_networks]
all_samples_features = np.vstack(vec_networks)


label_name = 'ABIDE_label_871.mat'
label = scio.loadmat(os.path.join(label_dir, label_name))
a = label['label']
a = a[0]
labels = np.zeros((number_samples, 2))
for i in range(number_samples):
    if a[i] == 1:
        labels[i,0] = 1
    else:
        labels[i,1] = 1

site_name = 'sites.mat'
site = scio.loadmat(os.path.join(label_dir, site_name))
all_sites = site['sites']
for i in range(len(all_sites)):
    b = all_sites[i]
    all_sites[i] = b.replace(' ', '')
unique_sites = np.unique(all_sites)
sites = list(all_sites)



######### population graph, only contains site information
graph = create_affinity_graph_from_scores(['sites'], subject_IDs, label_dir)


########################################### create selectors ##############################################################
dist_selectors = {}
dist_train = {}
dist_validation = {}
dist_test = {}
for i in range(k_fold):
    train_path = cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'train' + '\\'
    filenames = listdir(train_path)
    train_ind = []
    for j in filenames:
        index = int(j[:-4])
        train_ind.append(index)

    selector = feature_selection(all_samples_features, a, train_ind, new_number_features)
    dist_selectors[str(i+1)] = selector
    dist_train[str(i+1)] = train_ind
for i in range(k_fold):
    validation_path = cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'validation' + '\\'
    val_filenames = listdir(validation_path)
    val_ind = []
    test_path = cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'test' + '\\'
    test_filenames = listdir(test_path)
    test_ind = []
    for j in val_filenames:
        index = int(j[:-4])
        val_ind.append(index)
    for j in test_filenames:
        index = int(j[:-4])
        test_ind.append(index)
    dist_validation[str(i + 1)] = val_ind
    dist_test[str(i + 1)] = test_ind
########################################### Training ################################################
#####################################################################################################
#####################################################################################################
parser = argparse.ArgumentParser()
parser.add_argument('--ngpu', type=int, default=1)
# nEpochs is only used for the first 20 iterations
parser.add_argument('--nEpochs', type=int, default=500)
# second_nEpochs will be used in the second training part(iterations after 20)
parser.add_argument('--weight-decay', '--wd', default=1e-8, type=float,
                    metavar='W', help='weight decay (default: 1e-8)')
parser.add_argument('--no-cuda', action='store_true')
parser.add_argument('--seed', type=int, default=1)
# I define three kinds of optimizer, adam sgd and rmsprop, but adam performs the best
parser.add_argument('--opt', type=str, default='adam',
                    choices=('sgd', 'adam', 'rmsprop'))
parser.add_argument('--cheby_order_K', type=int, default=2)
parser.add_argument('--input_dimension', type=int, default=5000)
parser.add_argument('--hidden_dimension', type=int, default=32)
parser.add_argument('--output_dimension', type=int, default=2)
parser.add_argument('--dropout_rate', type=int, default=0.3)
#
args, unknown = parser.parse_known_args()
args.cuda = not args.no_cuda and torch.cuda.is_available()
weight_decay = args.weight_decay
torch.manual_seed(args.seed)
if args.cuda:
    torch.cuda.manual_seed(args.seed)


print("build GCN pytorch")
nll = False

model = ChebNet_one_dropout(in_c=args.input_dimension, hid_c=args.hidden_dimension, out_c=args.output_dimension, K=args.cheby_order_K, dropout_rate=args.dropout_rate)
gpu_ids = range(args.ngpu)
model = nn.parallel.DataParallel(model, device_ids=gpu_ids)



train = train_GCN
class_balance = False

print('  + Number of params: {}'.format(
    sum([p.data.nelement() for p in model.parameters()])))
if args.cuda:
    model = model.cuda()

if args.opt == 'sgd':
    optimizer = optim.SGD(model.parameters(),lr=1e-5, momentum=0.8)
elif args.opt == 'adam':
    optimizer = optim.Adam(model.parameters(), weight_decay=0.2)
elif args.opt == 'rmsprop':
    optimizer = optim.RMSprop(model.parameters(), lr=1e-3, weight_decay=weight_decay)

save_path = 'Choose a path where you want to store the models'
save_model(model,save_path, 'initial_model')

###################################### excel #########################################################
excel_root = "Choose a path where you want to record the accuracy and loss during the training validation and testing processes"
excel_name = 'AAL.xlsx'

workbook=load_workbook(os.path.join(excel_root, excel_name))
Sheet1 = workbook['Sheet1']

rows_each_section = 20
row_train_accuracy = 1
row_train_loss = 9
position = [row_train_accuracy, row_train_loss, rows_each_section]

Sheet1.cell(row=position[0], column=1, value='Train_accuracy')
Sheet1.cell(row=position[1], column=1, value='Train_loss')
Sheet1.cell(row=position[0] + position[2], column=1, value='Validation_accuracy')
Sheet1.cell(row=position[1] + position[2], column=1, value='Validation_loss')
Sheet1.cell(row=position[0] + position[2] * 2, column=1, value='Test_accuracy')
Sheet1.cell(row=position[1] + position[2] * 2, column=1, value='Test_loss')
Sheet1.cell(row=position[0] + position[2] * 3, column=1, value='Sensitivity')
Sheet1.cell(row=position[1] + position[2] * 3, column=1, value='Specificity')

################################################# Training ###################################################
for fold in range(1, k_fold+1):


    data = []
    fold_train_index = dist_train[str(fold)]
    fold_validation_index = dist_validation[str(fold)]
    fold_test_index = dist_test[str(fold)]
    fold_selector = dist_selectors[str(fold)]
    fold_data = fold_selector.transform(all_samples_features)
    #################################################################################
    n = fold_data.shape[0]
    num_edge = n * n
    edge_index = np.zeros([2, num_edge], dtype=np.int64)
    edgenet_input = np.zeros([num_edge, 1], dtype=np.float32)
    aff_score = np.zeros(num_edge, dtype=np.float32)
    flatten_ind = 0

    for i in range(n):
        for j in range(n):
            edge_index[:, flatten_ind] = [i, j]
            edgenet_input[flatten_ind] = graph[i,j]
            aff_score[flatten_ind] = graph[i,j]
            flatten_ind += 1
    keep_ind = np.where(aff_score > 0.99)[0]
    edge_index = edge_index[:, keep_ind]
    edgenet_input = edgenet_input[keep_ind]
    ######################################################################################
    data.append(fold_data)
    data.append(labels)
    data.append(edge_index)
    data.append(edgenet_input)

    Sheet1.cell(row=fold + position[0], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2], column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2] * 2, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2] * 2, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[0] + position[2] * 3, column=1, value='model_' + str(fold))
    Sheet1.cell(row=fold + position[1] + position[2] * 3, column=1, value='model_' + str(fold))

    for epoch in range(1, args.nEpochs + 1):
        if epoch == 1:
            model = load_model(model, save_path, 'initial_model')
            print('model: ' + str(epoch) + '_' + str(fold))
        else:
            model = load_model(model, save_path, str(epoch - 1) + '_' + str(fold))
            print('model: ' + str(epoch) + '_' + str(fold))

        adjust_opt(args.opt, optimizer, epoch)
        train(args, epoch, model, data, fold_train_index, fold_validation_index, fold_test_index, optimizer, fold, k_fold, save_path, Sheet1, position)

workbook.save(os.path.join(excel_root, excel_name))
