import os
import sys
import torch
import torch_geometric.datasets as GeoData
from torch_geometric.data import DataLoader, DataListLoader
from torch_geometric.nn.data_parallel import DataParallel
import random
import numpy as np

from opt import *
from EV_GCN import EV_GCN
from utils.metrics import accuracy, auc, prf
from dataloader import dataloader
from openpyxl import load_workbook
from os import listdir
import data.ABIDEParser as Reader
from sklearn.model_selection import StratifiedKFold


def save_model(net,path, name_net):
    path_net =  path + '/' + name_net + '.pth'
    torch.save(net.cpu().state_dict(), path_net)
    net.cuda()


if __name__ == '__main__':
    opt = OptInit().initialize()
########################################## k_fold = 5 ###############################################
#####################################################################################################
#####################################################################################################
    save_model_path = 'Input your own path (example: D:\EV_GCN\save_models\models)'

    subject_IDs = Reader.get_ids()
    labels_dist = Reader.get_subject_score(subject_IDs, score='DX_GROUP')
    sites_dist = Reader.get_subject_score(subject_IDs, score='SITE_ID')
    labels = list(labels_dist.values())
    sites = list(sites_dist.values())
    unique_sites = np.unique(list(sites_dist.values())).tolist()
    ############################################ 5 cross validation #################################################
    # based on labels and collection sites, to create a balanced dataset
    k_fold = 5
    dist_train = {}
    dist_validation = {}
    dist_test = {}
    for i in range(k_fold):
        dist_train[str(i + 1)] = []
        dist_validation[str(i + 1)] = []
        dist_test[str(i + 1)] = []

    for each_site in unique_sites:
        index_site = Reader.get_index(sites, each_site)
        label = np.zeros((len(index_site)))
        for i in range(len(index_site)):
            index = index_site[i]
            label[i] = int(labels[int(index)])

        sfolder = StratifiedKFold(n_splits=k_fold, random_state=0, shuffle=True)
        group = 1
        for train, validation in sfolder.split(index_site, label):
            for i in train:
                name = index_site[i]
                dist_train[str(group)].append(name)
            for j in validation:
                name = index_site[j]
                dist_validation[str(group)].append(name)
            group = group + 1
        group = 1
        for train, validation in sfolder.split(index_site, label):
            if group == 1:
                for j in validation:
                    name = index_site[j]
                    dist_train[str(k_fold)].remove(name)
                    dist_test[str(k_fold)].append(name)
            else:
                for j in validation:
                    name = index_site[j]
                    dist_train[str(group-1)].remove(name)
                    dist_test[str(group-1)].append(name)
            group = group + 1


    for i in range(k_fold):
        dist_train[str(i + 1)] = np.array(dist_train[str(i + 1)])
        dist_validation[str(i + 1)] = np.array(dist_validation[str(i + 1)])
        dist_test[str(i + 1)] = np.array(dist_test[str(i + 1)])

    ##############################################################################################################
    print('  Loading dataset ...')
    dl = dataloader()
    raw_features, y, nonimg = dl.load_data()
    # cv_splits = dl.data_split(k_fold)  #####

    corrects = np.zeros(k_fold, dtype=np.int32)
    accs = np.zeros(k_fold, dtype=np.float32)
    aucs = np.zeros(k_fold, dtype=np.float32)
    prfs = np.zeros([k_fold, 3], dtype=np.float32)

    ###################################### excel #########################################################
    # record the accuracy and loss during the training, validation and testing
    excel_root = "Input your own path"
    excel_name = 'test_2.xlsx'

    workbook = load_workbook(os.path.join(excel_root, excel_name))
    Sheet1 = workbook['Sheet1']

    rows_each_section = 30
    row_train_accuracy = 1
    row_train_loss = 14
    position = [row_train_accuracy, row_train_loss, rows_each_section]

    Sheet1.cell(row=position[0], column=1, value='Train_accuracy')
    Sheet1.cell(row=position[1], column=1, value='Train_loss')
    Sheet1.cell(row=position[0] + position[2], column=1, value='Validation_accuracy')
    Sheet1.cell(row=position[1] + position[2], column=1, value='Validation_loss')
    Sheet1.cell(row=position[0] + position[2] * 2, column=1, value='Test_accuracy')
    Sheet1.cell(row=position[1] + position[2] * 2, column=1, value='Test_loss')
    Sheet1.cell(row=position[0] + position[2] * 3, column=1, value='Sensitivity')
    Sheet1.cell(row=position[1] + position[2] * 3, column=1, value='Specificity')


    for fold in range(k_fold):

        Sheet1.cell(row=fold + 1 + position[0], column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[1], column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[0] + position[2], column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[1] + position[2], column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[0] + position[2] * 2, column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[1] + position[2] * 2, column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[0] + position[2] * 3, column=1, value='model_' + str(fold))
        Sheet1.cell(row=fold + 1 + position[1] + position[2] * 3, column=1, value='model_' + str(fold))

        print("\r\n========================== Fold {} ==========================".format(fold))
        train_ind = dist_train[str(fold+1)]
        val_ind = dist_validation[str(fold+1)]
        test_index = dist_test[str(fold+1)]
        print('  Constructing graph data...')
        # extract node features
        node_ftr = dl.get_node_features(train_ind)
        # get PAE inputs
        edge_index, edgenet_input = dl.get_PAE_inputs(nonimg)
        a = edgenet_input.mean(axis=0)
        edgenet_input = (edgenet_input - edgenet_input.mean(axis=0)) / edgenet_input.std(axis=0)

        # build network architecture
        model = EV_GCN(node_ftr.shape[1], opt.num_classes, opt.dropout, edge_dropout=opt.edropout, hgc=opt.hgc,
                       lg=opt.lg, edgenet_input_dim=2 * nonimg.shape[1]).to(opt.device)
        model = model.to(opt.device)

        # build loss, optimizer, metric
        loss_fn = torch.nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=opt.lr, weight_decay=opt.wd)
        features_cuda = torch.tensor(node_ftr, dtype=torch.float32).to(opt.device)
        edge_index = torch.tensor(edge_index, dtype=torch.long).to(opt.device)
        edgenet_input = torch.tensor(edgenet_input, dtype=torch.float32).to(opt.device)
        labels = torch.tensor(y, dtype=torch.long).to(opt.device)
        fold_model_path = opt.ckpt_path + "/fold{}.pth".format(fold)


        def train():
            print("  Number of training samples %d" % len(train_ind))
            print("  Start training...\r\n")
            for epoch in range(opt.num_iter):
                model.train()
                optimizer.zero_grad()

                with torch.set_grad_enabled(True):
                    node_logits, edge_weights = model(features_cuda, edge_index, edgenet_input)
                    loss = loss_fn(node_logits[train_ind], labels[train_ind])
                    loss.backward()
                    optimizer.step()
                correct_train, acc_train = accuracy(node_logits[train_ind].detach().cpu().numpy(), y[train_ind])

                Sheet1.cell(row=fold + 1 + position[0], column=epoch + 2, value=acc_train.item())
                Sheet1.cell(row=fold + 1 + position[1], column=epoch + 2, value=loss.item())
                print("Epoch: {},\tce loss: {:.5f},\ttrain acc: {:.5f}".format(epoch, loss.item(), acc_train.item()))
                save_model(model, save_model_path, str(epoch+1) + '_' + str(fold+1))


                model.eval()

                #### validation
                with torch.set_grad_enabled(False):
                    node_logits, _ = model(features_cuda, edge_index, edgenet_input)
                    loss = loss_fn(node_logits[val_ind], labels[val_ind])
                logits_val = node_logits[val_ind].detach().cpu().numpy()

                correct_val, acc_val = accuracy(logits_val, y[val_ind])
                Sheet1.cell(row=fold + 1 + position[0] + position[2], column=epoch + 2, value=acc_val.item())
                Sheet1.cell(row=fold + 1 + position[1] + position[2], column=epoch + 2, value=loss.item())

                #### testing
                with torch.set_grad_enabled(False):
                    node_logits, _ = model(features_cuda, edge_index, edgenet_input)
                    loss = loss_fn(node_logits[test_index], labels[test_index])
                logits_test = node_logits[test_index].detach().cpu().numpy()

                correct_test, acc_test = accuracy(logits_test, y[test_index])
                prf_test = prf(logits_test, y[test_index])
                Sheet1.cell(row=fold + 1 + position[0] + position[2] * 2, column=epoch + 2, value=acc_test.item())
                Sheet1.cell(row=fold + 1 + position[1] + position[2] * 2, column=epoch + 2, value=loss.item())
                Sheet1.cell(row=fold + 1 + position[0] + position[2] * 3, column=epoch + 2, value=prf_test[0])
                Sheet1.cell(row=fold + 1 + position[1] + position[2] * 3, column=epoch + 2, value=prf_test[1])

        if opt.train == 1:
            train()
    workbook.save(os.path.join(excel_root, excel_name))
