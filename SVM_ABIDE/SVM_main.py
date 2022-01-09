from load_data import *
from feature_selection import *
from sklearn import svm
from openpyxl import load_workbook
import os
import numpy as np
from os.path import join
from os import listdir
# %matplotlib inline



def calculate(labels, predictions):
    whole_number = 0
    TN = 0
    FP = 0
    FN = 0
    TP = 0
    for i in range(len(predictions)):
        if labels[i,0] == 1 and predictions[i] == 1:   # Autism
            TN = TN + 1
            whole_number = whole_number + 1
        elif labels[i,0] == 1 and predictions[i] == 0:
            FP = FP + 1
            whole_number = whole_number + 1
        elif labels[i,0] == 0 and predictions[i] == 1:
            FN = FN + 1
            whole_number = whole_number + 1
        elif labels[i,0] == 0 and predictions[i] == 0:  # TD
            TP = TP + 1
            whole_number = whole_number + 1
    accuracy = (TP + TN) / whole_number
    sensitivity = TP / (TP + FN)
    specificity = TN / (TN + FP)

    return accuracy, sensitivity, specificity

########################################### Load Data ###############################################
#####################################################################################################
#####################################################################################################
root_path = "Input the full path of \\download_data\\all  (all folder will be created by Data_download_prepare\\fetch_data.py)"
cross_validation_path = 'Input the full path of \\download_data\\cross validation  (cross validation folder will be created by Data_download_prepare\\fetch_data.py)'
label_dir = 'Input the full path of \\download_data\\measures  (measures folder will be created by Data_download_prepare\\fetch_data.py)'

k_fold = 5
image_size = [200, 200]
new_number_features = 5000

###################################### excel #########################################################
excel_root = "Choose a path where you want to record the accuracy and loss during the training validation and testing processes"
excel_name = 'cc200.xlsx'

workbook=load_workbook(os.path.join(excel_root, excel_name))
Sheet1 = workbook['Sheet1']
start_column = 1
start_row = 1
Sheet1.cell(row=start_row, column=start_column, value='validation')
Sheet1.cell(row=start_row, column=start_column+1, value='accuracy')
Sheet1.cell(row=start_row, column=start_column + 2, value='sensitivity')
Sheet1.cell(row=start_row, column=start_column + 3, value='specificity')
Sheet1.cell(row=start_row, column=start_column + 5, value='test')
Sheet1.cell(row=start_row, column=start_column+6, value='accuracy')
Sheet1.cell(row=start_row, column=start_column + 7, value='sensitivity')
Sheet1.cell(row=start_row, column=start_column + 8, value='specificity')

######################################################################################################
label_name = 'ABIDE_label_871.mat'
label = scio.loadmat(os.path.join(label_dir, label_name))
a = label['label']
a = a[0]
filenames = listdir(root_path)
number = len(filenames)
length = image_size[0] * (image_size[1]-1) / 2
group_images = np.zeros((number, int(length)))

for index in range(number):
    image_name = str(filenames[index])
    image = scio.loadmat(os.path.join(root_path, image_name))
    img = image['connectivity']
    group_images[index,:] = flatten_one(length, img)


dist_selectors = {}
for i in range(k_fold):
    train_path = cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'train' + '\\'
    filenames = listdir(train_path)
    train_ind = []
    for j in filenames:
        index = int(j[:-4])
        train_ind.append(index)

    selector = feature_selection(group_images, a, train_ind, new_number_features)
    dist_selectors[str(i + 1)] = selector


dist_train = {}
dist_validation = {}
dist_test = {}
for i in range(k_fold):
    fold_list = []
    for j in range(k_fold):
        fold_list.append('set' + str(j))
    del fold_list[i]
    dist_validation['validation_images_'+ str(i + 1)], dist_validation['validation_labels_'+ str(i + 1)]= \
        load_data_SVM(join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'validation'),
                         image_size,
                         join(label_dir),
                         label_name,
                         selector = dist_selectors[str(i+1)],
                         new_number_features = new_number_features)

    dist_train['train_images_'+ str(i + 1)], dist_train['train_labels_'+ str(i + 1)]= \
        load_data_SVM(join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'train'),
                         image_size,
                         join(label_dir),
                         label_name,
                         selector = dist_selectors[str(i+1)],
                         new_number_features = new_number_features)

    dist_test['test_images_'+ str(i + 1)], dist_test['test_labels_'+ str(i + 1)]= \
        load_data_SVM(join(cross_validation_path + '\\' + 'group' + str(i) + '\\' + 'test'),
                         image_size,
                         join(label_dir),
                         label_name,
                         selector = dist_selectors[str(i+1)],
                         new_number_features = new_number_features)


###### SVM model
model = svm.SVC(kernel='rbf', C=3)

# ################################################# Training ###################################################
average_validation = 0
sensitivity_val = 0
specificity_val = 0
average_test = 0
sensitivity_test = 0
specificity_test = 0
for fold in range(1, k_fold+1):
    Sheet1.cell(row=start_row + fold, column=start_column, value='model_' + str(fold))
    Sheet1.cell(row=start_row + fold, column=start_column + 5, value='model_' + str(fold))



    print('Model_'+ str(fold))
    train_images = dist_train['train_images_' + str(fold)]
    train_labels = dist_train['train_labels_' + str(fold)]
    validation_images = dist_validation['validation_images_' + str(fold)]
    validation_labels = dist_validation['validation_labels_' + str(fold)]
    test_images = dist_test['test_images_' + str(fold)]
    test_labels = dist_test['test_labels_' + str(fold)]

    model.fit(train_images,train_labels.ravel())

    val_results = model.predict(validation_images)
    test_results = model.predict(test_images)

    val_accuracy, val_sensitivity, val_specificity = calculate(validation_labels, val_results)
    test_accuracy, test_sensitivity, test_specificity = calculate(test_labels, test_results)

    average_validation = val_accuracy + average_validation
    sensitivity_val = val_sensitivity + sensitivity_val
    specificity_val = val_specificity + specificity_val
    average_test = average_test + test_accuracy
    sensitivity_test = test_sensitivity + sensitivity_test
    specificity_test = test_specificity + specificity_test
    print('validation' + str(
            fold) + ': accuracy : {:.8f}\t sensitivity : {:.8f}\t specificity : {:.8f}\t'.format(
            val_accuracy, val_sensitivity, val_specificity))
    print('test' + str(
        fold) + ': accuracy : {:.8f}\t sensitivity : {:.8f}\t specificity : {:.8f}\t'.format(
        test_accuracy, test_sensitivity, test_specificity))

    Sheet1.cell(row=start_row+ fold, column=start_column + 1, value=val_accuracy)
    Sheet1.cell(row=start_row+ fold, column=start_column + 2, value=val_sensitivity)
    Sheet1.cell(row=start_row+ fold, column=start_column + 3, value=val_specificity)

    Sheet1.cell(row=start_row+ fold, column=start_column + 6, value=test_accuracy)
    Sheet1.cell(row=start_row+ fold, column=start_column + 7, value=test_sensitivity)
    Sheet1.cell(row=start_row+ fold, column=start_column + 8, value=test_specificity)



average_validation = average_validation / k_fold
sensitivity_val = sensitivity_val / k_fold
specificity_val = specificity_val / k_fold

average_test = average_test / k_fold
sensitivity_test = sensitivity_test / k_fold
specificity_test = specificity_test / k_fold

Sheet1.cell(row=start_row + k_fold+1, column=start_column, value='average')
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 1, value=average_validation)
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 2, value=sensitivity_val)
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 3, value=specificity_val)

Sheet1.cell(row=start_row + k_fold+1, column=start_column + 5, value='average')
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 6, value=average_test)
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 7, value=sensitivity_test)
Sheet1.cell(row=start_row + k_fold+1, column=start_column + 8, value=specificity_test)
print('\n')
print('The average validation is : ', average_validation)
print('The average test is : ', average_test)

workbook.save(os.path.join(excel_root, excel_name))
